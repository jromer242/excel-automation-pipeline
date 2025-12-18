"""
Methods to Edit a Single Worksheet Without Affecting Others
"""

import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows


# ============================================================================
# METHOD 1: Using openpyxl (RECOMMENDED - Most Control)
# ============================================================================

def edit_worksheet_openpyxl(file_path, sheet_name, modifications):
    """
    Edit a specific worksheet using openpyxl.
    This preserves ALL other sheets, formatting, formulas, etc.

    Args:
        file_path: Path to Excel file
        sheet_name: Name of sheet to edit
        modifications: Function that takes worksheet and modifies it
    """
    # Load the entire workbook
    workbook = openpyxl.load_workbook(file_path)

    # Get the specific worksheet
    worksheet = workbook[sheet_name]

    # Make your modifications
    modifications(worksheet)

    # Save back to the same file (or a new file)
    workbook.save(file_path)
    workbook.close()
    print(f"✓ Updated '{sheet_name}' in {file_path}")


def example_openpyxl_modifications():
    """Example: Update specific cells in a worksheet"""

    # Create a sample workbook with multiple sheets
    wb = openpyxl.Workbook()

    # Sheet 1: Sales Data
    ws1 = wb.active
    ws1.title = "Sales"
    ws1.append(["Product", "Q1", "Q2", "Q3", "Q4"])
    ws1.append(["Widget A", 100, 150, 200, 175])
    ws1.append(["Widget B", 80, 90, 110, 95])

    # Sheet 2: Inventory
    ws2 = wb.create_sheet("Inventory")
    ws2.append(["Product", "Stock", "Reorder Point"])
    ws2.append(["Widget A", 50, 20])
    ws2.append(["Widget B", 15, 25])

    # Sheet 3: Customers
    ws3 = wb.create_sheet("Customers")
    ws3.append(["Customer", "Region", "Type"])
    ws3.append(["Acme Corp", "North", "Wholesale"])

    wb.save("sample_workbook.xlsx")
    print("✓ Created sample_workbook.xlsx with 3 sheets")

    # Now edit ONLY the Sales sheet
    def update_sales(ws):
        # Update Q4 sales for Widget A
        ws['E2'] = 250  # Direct cell reference

        # Add a new row
        ws.append(["Widget C", 60, 70, 80, 90])

        # Update a cell by row/column index (row 3, col 3 = B3)
        ws.cell(row=3, column=3, value=100)

    edit_worksheet_openpyxl("sample_workbook.xlsx", "Sales", update_sales)
    print("✓ Sales sheet updated (Inventory and Customers untouched)")


# ============================================================================
# METHOD 2: Using pandas with openpyxl engine (Good for DataFrame operations)
# ============================================================================

def edit_worksheet_pandas_safe(file_path, sheet_name, dataframe_modifier):
    """
    Edit a worksheet using pandas while preserving other sheets.

    Args:
        file_path: Path to Excel file
        sheet_name: Name of sheet to edit
        dataframe_modifier: Function that takes a DataFrame and returns modified DataFrame
    """
    # Read the specific sheet into a DataFrame
    df = pd.read_excel(file_path, sheet_name=sheet_name)

    # Modify the DataFrame
    modified_df = dataframe_modifier(df)

    # Read ALL sheet names
    with pd.ExcelFile(file_path) as xls:
        sheet_names = xls.sheet_names

    # Write back using ExcelWriter in 'openpyxl' mode
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        modified_df.to_excel(writer, sheet_name=sheet_name, index=False)

    print(f"✓ Updated '{sheet_name}' using pandas")


def example_pandas_modifications():
    """Example: Modify data using pandas operations"""

    # Create sample file
    with pd.ExcelWriter("sales_data.xlsx", engine='openpyxl') as writer:
        # Sheet 1
        pd.DataFrame({
            'Product': ['A', 'B', 'C'],
            'Price': [10, 20, 30],
            'Stock': [100, 50, 75]
        }).to_excel(writer, sheet_name='Products', index=False)

        # Sheet 2
        pd.DataFrame({
            'Date': ['2024-01-01', '2024-01-02'],
            'Sales': [1000, 1500]
        }).to_excel(writer, sheet_name='Daily_Sales', index=False)

    print("✓ Created sales_data.xlsx with 2 sheets")

    # Modify only the Products sheet
    def modify_products(df):
        # Add a 10% markup to prices
        df['Price'] = df['Price'] * 1.10

        # Add a new column
        df['Value'] = df['Price'] * df['Stock']

        # Filter out low stock items
        df = df[df['Stock'] > 60]

        return df

    edit_worksheet_pandas_safe("sales_data.xlsx", "Products", modify_products)
    print("✓ Products sheet updated (Daily_Sales untouched)")


# ============================================================================
# METHOD 3: Read All, Modify One, Write All (Simple but less efficient)
# ============================================================================

def edit_worksheet_read_all(file_path, sheet_name, dataframe_modifier):
    """
    Read all sheets, modify one, write all back.
    Simple but loses some formatting.
    """
    # Read ALL sheets into a dictionary
    all_sheets = pd.read_excel(file_path, sheet_name=None)

    # Modify the specific sheet
    if sheet_name in all_sheets:
        all_sheets[sheet_name] = dataframe_modifier(all_sheets[sheet_name])
    else:
        print(f"Warning: Sheet '{sheet_name}' not found!")
        return

    # Write all sheets back
    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        for sheet, df in all_sheets.items():
            df.to_excel(writer, sheet_name=sheet, index=False)

    print(f"✓ Updated '{sheet_name}' (rewrote all sheets)")


# ============================================================================
# METHOD 4: Advanced - Replace DataFrame in existing worksheet with openpyxl
# ============================================================================

def replace_worksheet_data_openpyxl(file_path, sheet_name, new_df, start_row=1, start_col=1):
    """
    Replace data in a worksheet with a DataFrame while preserving workbook structure.
    Great for updating data while keeping formatting in other cells.

    Args:
        file_path: Path to Excel file
        sheet_name: Sheet to update
        new_df: New DataFrame to write
        start_row: Starting row (1-indexed)
        start_col: Starting column (1-indexed)
    """
    wb = openpyxl.load_workbook(file_path)
    ws = wb[sheet_name]

    # Clear existing data in the range (optional)
    # ws.delete_rows(start_row, ws.max_row)

    # Write DataFrame to worksheet
    for r_idx, row in enumerate(dataframe_to_rows(new_df, index=False, header=True), start=start_row):
        for c_idx, value in enumerate(row, start=start_col):
            ws.cell(row=r_idx, column=c_idx, value=value)

    wb.save(file_path)
    wb.close()
    print(f"✓ Replaced data in '{sheet_name}' starting at row {start_row}, col {start_col}")


# ============================================================================
# PRACTICAL EXAMPLE: Monthly Report Update
# ============================================================================

def practical_example():
    """Real-world scenario: Update monthly sales without touching other sheets"""

    # Create a workbook with multiple sheets (simulating an existing report)
    wb = openpyxl.Workbook()

    # Summary sheet (should NOT be touched)
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.append(["Metric", "Value"])
    ws_summary.append(["Total Sales", "=SUM(Monthly_Data!B:B)"])
    ws_summary.append(["Report Date", "2024-12-17"])

    # Monthly data sheet (WILL be updated)
    ws_monthly = wb.create_sheet("Monthly_Data")
    ws_monthly.append(["Date", "Sales", "Units"])
    ws_monthly.append(["2024-01", 10000, 100])
    ws_monthly.append(["2024-02", 12000, 120])

    # Configuration sheet (should NOT be touched)
    ws_config = wb.create_sheet("Config")
    ws_config.append(["Setting", "Value"])
    ws_config.append(["Currency", "USD"])
    ws_config.append(["Region", "North America"])

    wb.save("monthly_report.xlsx")
    print("✓ Created monthly_report.xlsx")

    # Now update ONLY the Monthly_Data sheet with new data
    print("\n--- Updating Monthly_Data sheet ---")

    # Read current monthly data
    df = pd.read_excel("monthly_report.xlsx", sheet_name="Monthly_Data")

    # Add new month
    new_row = pd.DataFrame({
        'Date': ['2024-03'],
        'Sales': [15000],
        'Units': [150]
    })
    df = pd.concat([df, new_row], ignore_index=True)

    # Write back using mode='a' to append/replace
    with pd.ExcelWriter("monthly_report.xlsx", engine='openpyxl', mode='a',
                        if_sheet_exists='replace') as writer:
        df.to_excel(writer, sheet_name="Monthly_Data", index=False)

    print("✓ Added March data to Monthly_Data")
    print("✓ Summary and Config sheets remain unchanged")
    print("✓ Formulas in Summary sheet still work!")


# ============================================================================
# COMPARISON & RECOMMENDATIONS
# ============================================================================

def print_recommendations():
    """Guide on which method to use"""
    print("\n" + "=" * 70)
    print("WHICH METHOD SHOULD YOU USE?")
    print("=" * 70)

    recommendations = """
    📌 METHOD 1 (openpyxl) - BEST FOR:
       • Updating specific cells
       • Preserving ALL formatting, formulas, charts
       • Maximum control over Excel features
       • When you need to keep existing structure intact

    📊 METHOD 2 (pandas + openpyxl mode='a') - BEST FOR:
       • DataFrame-style operations (filter, group, calculate)
       • Replacing entire sheet data
       • When other sheets must stay untouched
       • Balanced approach for data analysts

    🔄 METHOD 3 (Read All, Write All) - BEST FOR:
       • Simple files with few sheets
       • When formatting doesn't matter
       • Quick scripts where efficiency isn't critical

    🎯 METHOD 4 (DataFrame to openpyxl) - BEST FOR:
       • Updating a data range within a formatted sheet
       • Keeping headers/footers/formatting separate from data
       • Advanced reporting scenarios

    ⚠️  IMPORTANT NOTES:
       • Always test on a copy first!
       • pandas may lose some Excel formatting
       • openpyxl preserves everything but is more verbose
       • Use mode='a' with if_sheet_exists='replace' for pandas
    """
    print(recommendations)


# ============================================================================
# RUN EXAMPLES
# ============================================================================

if __name__ == "__main__":
    print("SAFE EXCEL WORKSHEET EDITING EXAMPLES\n")

    # Example 1: openpyxl method
    print("\n[Example 1: openpyxl - Maximum Control]")
    example_openpyxl_modifications()

    # Example 2: pandas method
    print("\n[Example 2: pandas - DataFrame Operations]")
    example_pandas_modifications()

    # Example 3: Real-world scenario
    print("\n[Example 3: Real-World Monthly Report Update]")
    practical_example()

    # Show recommendations
    print_recommendations()

    print("\n" + "=" * 70)
    print("All examples complete! Check the generated Excel files.")
    print("=" * 70)